from re import A
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np

df = pd.read_excel(r'C:\Users\aalon\OneDrive\Desktop\python\crawler\targets.xlsx')
df['Redfin'] = np.nan

book = openpyxl.load_workbook(r'C:\Users\aalon\OneDrive\Desktop\python\crawler\targets.xlsx')
sheet = book.active


PATH = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(PATH)
driver.get("https://www.redfin.com/?&utm_source=google&utm_medium=ppc&utm_campaign=1022779&utm_term=aud-849690325664:kwd-844252101&utm_content=449276932284&adgid=105971835338&gclid=CjwKCAjwm8WZBhBUEiwA178UnPIRxF4Aqv_sZ-bR1aQ-EdsVenwcFghiyKcMnN9rq3b8tlOuekhyTxoCS8IQAvD_BwE&gclsrc=aw.ds")
print(driver.title)


for row in range(1, sheet.max_row):
    print('***********************')
    property = sheet.cell(row=row+1, column=22).value
    property = property + ' utah'
    search = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, 'search-box-input'))
    )

    search.send_keys(property)
    time.sleep(2)
    search.send_keys(Keys.RETURN)
    time.sleep(2)
    search.submit

    try:
        try:
            # This code executes when there's a property match. 
            market_status = driver.find_element(By.CSS_SELECTOR, '.secondary-stats-banner .off-market, .secondary-stats-banner .recently-sold')
            print(property)
            print(market_status.text)
            df.at[row - 1, 'Redfin'] = market_status.text

        except:
            # If no match is found . . . 
            print(f'No match found for {property}.')
            df.at[row - 1, 'Redfin'] = 'Not_found'
            
        try:
            results = driver.find_element(By.CSS_SELECTOR, '.home-main-stats-variant')
            print(results.text)
        except:
            print('')

        try:
            # If There's a related property found with a similar address, this should select the first option
            related_search_results = driver.find_element(By.CSS_SELECTOR, '.Dialog>.cell>.guts>.header>h3')
            did_you_mean = driver.find_element(By.CSS_SELECTOR, '.Dialog.searchDisambigDialog .content .ExpandedResults .expanded-section .expanded-row-content .item-row-show-sections').click()
            print(related_search_results.text)
        
        except:
            pass
        
        try:
            clear_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'search-box-input'))
            )
            clear_button.click()
        except:
            continue

        try:
            on_market = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.MortgageQuoteEntryPoint a'))
            )
            print('Property is on the market!')
        except:
            pass
    except:
        print('Property is for rent, not for sale.')
            

print(df[['prop_locat', 'Redfin']])
df.to_csv (r'C:\Users\aalon\OneDrive\Desktop\python\crawler\redfin_df.csv', index = None, header=True) 

    # search.clear()
#     try:
#         try:
#             content = WebDriverWait(driver, 10).until(
#                 EC.presence_of_element_located((By.CSS_SELECTOR, 'p.nearby-results__text strong'))
#             )
#             if content.text == 'Your search did not match any properties, but here are some nearby.':
#                 results = content.text
#                 df.at[row - 1, 'Market_Status'] = 'Not for sale'
#                 print(f"{property} is not for sale.")
#                 content = driver.find_element(By.CSS_SELECTOR, '.main-header .logo').click()
#             else:
#                 print('Oooops')

#         except:
#             content = WebDriverWait(driver, 10).until(
#             EC.presence_of_element_located((By.CSS_SELECTOR, '.profile-hero-wrapper .profile-hero-title'))
#             )
#             df.at[row - 1, 'Market_Status'] = 'FOR SALE'
#             print(content.text)
#             print(f"{property} is for sale!.")
#             content = driver.find_element(By.CSS_SELECTOR, '.ldp-header .ln-logo .logo-default').click()

#         try:
#             content = driver.find_element(By.CSS_SELECTOR, '.ldp-header .ln-logo').click()
#             print('second TRY')

#         except:
#             print('second EXCEPT')
#     except:
#         print('Everything you tried failed')
#         content = driver.find_element(By.CSS_SELECTOR, '.main-header .logo img.logo-default, .main-header.default .logo img.logo-default').click()

# print(df[['prop_locat', 'Market_Status']])

# df.to_csv (r'C:\Users\aalon\OneDrive\Desktop\python\crawler\export_dataframe.csv', index = None, header=True) 
    


