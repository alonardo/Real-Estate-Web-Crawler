from re import A
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np

df = pd.read_excel(r'C:\Users\aalon\OneDrive\Desktop\python\crawler\targets.xlsx')
df['Market_Status'] = np.nan

book = openpyxl.load_workbook(r'C:\Users\aalon\OneDrive\Desktop\python\crawler\targets.xlsx')
sheet = book.active

PATH = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(PATH)
driver.get("https://www.loopnet.com/")
print(driver.title)

# for row in range(1, sheet.max_row + 1):

for row in range(1, sheet.max_row):
    print(df[['prop_locat', 'Market_Status']])
    property = sheet.cell(row=row+1, column=22).value
    property = property + ' utah'

    search = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.NAME, 'geography'))
    )

    search.send_keys(property)
    search.send_keys(Keys.RETURN)
    search.submit
    try:
        try:
            content = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'p.nearby-results__text strong'))
            )
            if content.text == 'Your search did not match any properties, but here are some nearby.':
                results = content.text
                df.at[row - 1, 'Market_Status'] = 'Not for sale'
                print(f"{property} is not for sale.")
                content = driver.find_element(By.CSS_SELECTOR, '.main-header .logo').click()
            else:
                print('Oooops')

        except:
            content = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.profile-hero-wrapper .profile-hero-title'))
            )
            df.at[row - 1, 'Market_Status'] = 'FOR SALE'
            print(content.text)
            print(f"{property} is for sale!.")
            content = driver.find_element(By.CSS_SELECTOR, '.ldp-header .ln-logo .logo-default').click()

        try:
            content = driver.find_element(By.CSS_SELECTOR, '.ldp-header .ln-logo').click()
            print('second TRY')

        except:
            print('second EXCEPT')
    except:
        print('Everything you tried failed')
        content = driver.find_element(By.CSS_SELECTOR, '.main-header .logo img.logo-default, .main-header.default .logo img.logo-default').click()

print(df[['prop_locat', 'Market_Status']])

df.to_csv (r'C:\Users\aalon\OneDrive\Desktop\python\crawler\export_dataframe.csv', index = None, header=True) 
    


